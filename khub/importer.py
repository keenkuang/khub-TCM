"""老问诊系统数据导入器。

支持 Excel (.xlsx) 和 HTML 表格格式，自动识别中文字段名，
将问诊记录导入 khub 的 患者/病历/问诊 模块。

用法::

    from .importer import LegacyImporter
    imp = LegacyImporter(store)
    result = imp.import_excel("patients.xlsx")
    print(f"导入完成：{result['imported']} 条")
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
from typing import Optional

from .clinical.patients import add_patient
from .clinical.records import add_record
from .clinical.consultations import add_consultation
from .db import Store

logger = logging.getLogger("khub.importer")

# 字段名映射：中文 → 内部字段名
_PATIENT_FIELDS = {
    "患者id": "pid", "患者编号": "pid", "编号": "pid", "id": "pid", "序号": "pid",
    "姓名": "name", "患者姓名": "name", "病人姓名": "name",
    "性别": "gender",
    "出生日期": "born", "出生": "born", "生日": "born", "生辰": "born",
    "真实出生日期": "born",
    "电话": "phone", "手机": "phone", "联系电话": "phone", "手机号": "phone",
}

_CONSULT_FIELDS = {
    "就诊日期": "visit_date", "日期": "visit_date", "就诊时间": "visit_date",
    "提交时间": "visit_date",
    "主诉": "chief_complaint",
    "我第一想解决的问题": "chief_complaint",
    "我第二想解决的问题": "chief2",
    "我第三想解决的问题": "chief3",
    "舌苔": "tongue_pulse", "舌象": "tongue_pulse", "舌诊": "tongue_pulse",
    "脉象": "tongue_pulse", "脉诊": "tongue_pulse",
    "诊断": "diagnosis", "西医诊断": "diagnosis",
    "辨证": "differentiation", "证型": "differentiation", "中医辨证": "differentiation",
    "处方": "prescription", "方剂": "prescription", "中药方": "prescription",
    "治疗方案": "plan", "治疗计划": "plan", "方案": "plan",
    "胃口": "appetite",
    "备注": "note", "附注": "note", "说明": "note",
    "请填写其他需要补充的": "note",
    "最后还有什么觉得需要补充说明的情况": "note",
    "睡眠有哪些异常状况": "sleep",
    "精神与体力状况": "energy",
    "诊别": "visit_type",
    "我要看的医生是": "doctor",
}


class LegacyImporter:
    """老问诊系统数据导入器。"""

    def __init__(self, store: Store):
        self.store = store
        self._stats = {"patients": 0, "records": 0, "consultations": 0, "errors": []}

    def import_excel(self, path: str, sheet: str | int = 0,
                     dry_run: bool = False) -> dict:
        """从 Excel 文件导入问诊数据。

        Args:
            path: .xlsx 文件路径。
            sheet: 工作表名或索引（默认第一张）。
            dry_run: True 时只解析不写入。

        Returns:
            {"patients": N, "records": N, "consultations": N, "errors": [...]}
        """
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return self._stats

        # 跳过标题行（内容包含"反馈数据"、"导出"等关键字，或行中大部分为空）
        data_start = 0
        for i, row in enumerate(rows):
            first_cell = str(row[0] or "").strip() if row else ""
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty > 3 and "反馈数据" not in first_cell and "导出" not in first_cell:
                data_start = i
                break

        headers = [str(h or "").strip() for h in rows[data_start]]
        col_map = self._build_col_map(headers)

        for row in rows[data_start + 1:]:
            # 跳过空行/页脚行（包含 CRM 水印、无关键数据）
            first_cell = str(row[0] or "").strip() if row else ""
            if not row or all(c is None or not str(c).strip() for c in row):
                continue
            row_text = " ".join(str(c or "") for c in row)
            if "mikecrm" in row_text.lower() or "麦客CRM" in row_text or "技术支持" in row_text:
                continue
            # 至少需要一个有效字段（姓名、序号、或主诉）
            name = str(row[col_map.get("name", -1)] or "").strip() if col_map.get("name", -1) >= 0 else ""
            pid_from_row = str(row[col_map.get("pid", -1)] or "").strip() if col_map.get("pid", -1) >= 0 else ""
            if not name and not pid_from_row:
                continue
            try:
                self._import_row(col_map, row, dry_run)
            except Exception as exc:
                logger.warning("导入行失败: %s", exc)
                self._stats["errors"].append({"row": row, "error": str(exc)})

        return self._stats

    def import_html(self, path_or_text: str, dry_run: bool = False) -> dict:
        """从 HTML 文件（含 <table>）导入问诊数据。"""
        from html.parser import HTMLParser

        class TableParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.tables = []
                self._cur_table = []
                self._cur_row = []
                self._cur_cell = ""
                self._in_td = False
                self._in_tr = False

            def handle_starttag(self, tag, attrs):
                if tag == "tr":
                    self._cur_row = []
                    self._in_tr = True
                elif tag in ("td", "th"):
                    self._cur_cell = ""
                    self._in_td = True

            def handle_endtag(self, tag):
                if tag == "tr" and self._in_tr:
                    self._cur_table.append(list(self._cur_row))
                    self._in_tr = False
                elif tag in ("td", "th") and self._in_td:
                    self._cur_row.append(self._cur_cell.strip())
                    self._in_td = False
                elif tag == "table" and self._cur_table:
                    self.tables.append(self._cur_table)
                    self._cur_table = []

            def handle_data(self, data):
                if self._in_td:
                    self._cur_cell += data

        content = path_or_text
        if os.path.isfile(path_or_text):
            with open(path_or_text, encoding="utf-8", errors="replace") as f:
                content = f.read()

        parser = TableParser()
        parser.feed(content)
        if not parser.tables:
            raise ValueError("未在文件/内容中找到 HTML 表格")

        for table in parser.tables:
            if not table:
                continue
            headers = [str(h or "").strip() for h in table[0]]
            col_map = self._build_col_map(headers)
            for row in table[1:]:
                try:
                    self._import_row(col_map, row, dry_run)
                except Exception as exc:
                    logger.warning("导入行失败: %s", exc)
                    self._stats["errors"].append({"row": row, "error": str(exc)})

        return self._stats

    def _build_col_map(self, headers: list[str]) -> dict[str, int]:
        """根据表头建立字段名→列索引的映射。"""
        col_map = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower().replace(" ", "").replace("　", "")
            # 精确匹配
            if h in _PATIENT_FIELDS:
                col_map[_PATIENT_FIELDS[h]] = idx
            elif h in _CONSULT_FIELDS:
                col_map[_CONSULT_FIELDS[h]] = idx
            # 模糊匹配
            else:
                for zh, en in {**_PATIENT_FIELDS, **_CONSULT_FIELDS}.items():
                    if zh in h_lower or h_lower in zh.lower():
                        col_map.setdefault(en, idx)
        return col_map

    def _import_row(self, col_map: dict, row: tuple, dry_run: bool):
        """导入单行数据。"""
        pid = self._val(row, col_map, "pid") or self._gen_pid()

        # 患者信息
        name = self._val(row, col_map, "name")
        gender = self._val(row, col_map, "gender")
        born = self._val(row, col_map, "born")
        phone = self._val(row, col_map, "phone")

        if name:
            if not dry_run:
                add_patient(self.store, pid, name, gender=gender, born=born)
            self._stats["patients"] += 1

        # 合并多项主诉
        chief = self._val(row, col_map, "chief_complaint")
        chief2 = self._val(row, col_map, "chief2")
        chief3 = self._val(row, col_map, "chief3")
        chief_parts = [c for c in (chief, chief2, chief3) if c]
        combined_chief = "；".join(chief_parts) if chief_parts else ""

        # 问诊记录（没有主诉的标记为复诊）
        if not combined_chief:
            visit_type = self._val(row, col_map, "visit_type") or ""
            doctor = self._val(row, col_map, "doctor") or ""
            combined_chief = f"复诊" + (f"（医生：{doctor}）" if doctor else "")
        differentiation = self._val(row, col_map, "differentiation")
        plan = self._val(row, col_map, "plan")
        appetite = self._val(row, col_map, "appetite")

        if not dry_run:
            add_consultation(self.store, pid,
                             chief_complaint=combined_chief,
                             tongue_pulse=self._val(row, col_map, "tongue_pulse") or "",
                             differentiation=differentiation or "",
                             plan=plan or "")
        self._stats["consultations"] += 1

        # 病历记录（有诊断/处方则写，否则用主诉作为诊断参考）
        diagnosis = self._val(row, col_map, "diagnosis")
        prescription = self._val(row, col_map, "prescription")
        note = self._val(row, col_map, "note") or combined_chief

        if not dry_run:
            add_record(self.store, pid,
                       diagnosis=diagnosis or ("（待补充）" if not diagnosis else ""),
                       prescription=prescription or "",
                       note=note or "")
        self._stats["records"] += 1

    @staticmethod
    def _val(row: tuple, col_map: dict, key: str) -> str:
        """从行中取指定字段的值。"""
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        if v is None:
            return ""
        return str(v).strip()

    @staticmethod
    def _gen_pid() -> str:
        return f"legacy-{int(time.time() * 1000)}"
